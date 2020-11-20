#!/usr/bin/env python3
# _*_ coding: utf-8 _*_
# @Author: ChenAnming
# @Time: 2020/11/4 0004 13:58
# @File: yamlset.py
# @Poject: Work_Project

import shutil
import openpyxl
from YouShuYun_API.config.config import CF
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from YouShuYun_API.common.variable import is_vars


class YamlSet:
    """Excel配置"""

    def __init__(self):
        shutil.copyfile(is_vars.get('excel_input'), is_vars.get('excel_output'))
        self.path = is_vars.get('excel_output')
        self.wb = openpyxl.load_workbook(self.path)
        self.table = self.wb.active

    def get_cases(self, min_row=2):
        """获取用例"""
        all_cases = []
        for row in self.table.iter_rows(min_row=min_row):
            all_cases.append((self.table.cell(min_row, CF.NAME + 1).value,
                              min_row, [cell.value for cell in row]))
            min_row += 1
        return all_cases

    def write_color(self, row_n, col_n, color=CF.COLOR_FAILED):
        """写入颜色"""
        cell = self.table.cell(row_n, col_n + 1)
        fill = PatternFill("solid", fgColor=color)
        cell.fill = fill

    def write_results(self, row_n, col_n, value, color=True):
        """写入结果"""
        cell = self.table.cell(row_n, col_n + 1)
        cell.value = value
        font = Font(name=CF.FONT_SET, size=CF.FONT_SIZE)
        cell.font = font
        if color:
            if value.lower() in ("fail", 'failed'):
                fill = PatternFill("solid", fgColor=CF.COLOR_FAILED)
                cell.fill = fill
            elif value.lower() in ("pass", "ok"):
                fill = PatternFill("solid", fgColor=CF.COLOR_PASSED)
                cell.fill = fill
        self.wb.save(self.path)


yaml_set = YamlSet()
if __name__ == '__main__':
    print(yaml_set.get_cases())